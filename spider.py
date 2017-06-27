import time
import urllib
from openpyxl import Workbook
from urllib import parse, request, error

import numpy as np
from bs4 import BeautifulSoup

hds = [{'User-Agent': 'Mozilla/5.0 (Windows; U; Windows NT 6.1; en-US; rv:1.9.1.6) Gecko/20091201 Firefox/3.5.6'},
       {
           'User-Agent': 'Mozilla/5.0 (Windows NT 6.2) AppleWebKit/535.11 (KHTML, like Gecko) Chrome/17.0.963.12 Safari/535.11'},
       {'User-Agent': 'Mozilla/5.0 (compatible; MSIE 10.0; Windows NT 6.2; Trident/6.0)'}]


def book_spider(book_tag):
    page_num = 0
    book_list = []
    try_times = 0
    while True:
        url = 'https://www.douban.com/tag/' + parse.quote_plus(book_tag) + '/book?start=' + str(page_num * 15)
        time.sleep(np.random.rand() * 5)
        try:
            req = request.Request(url, headers=hds[page_num % len(hds)])
            source_code = request.urlopen(req)
        except urllib.error.HTTPError as e:
            print(1, e)
            continue

        soup = BeautifulSoup(source_code, "lxml")
        list_soup = soup.find('div', {'class': 'mod book-list'})
        if not list_soup:
            if try_times < 200:
                try_times += 1
                continue
            else:
                break

        for book_info in list_soup.findAll('dd'):
            title = book_info.find('a', {'class': 'title'}).string.strip()
            desc = book_info.find('div', {'class': 'desc'}).string.strip()
            desc_list = desc.split('/')
            for i in range(len(desc_list)):
                desc_list[i].strip()
            book_url = book_info.find('a', {'class': 'title'}).get('href')

            try:
                author_info = 'Author: ' + ' / '.join(desc_list[0:-3])
            except:
                author_info = 'Author Unknown'

            try:
                pub_info = 'Publication: ' + ' / '.join(desc_list[-3:])
            except:
                pub_info = 'Publication Unknown'

            try:
                rating = book_info.find('span', {'class': 'rating_nums'}).string.strip()
            except:
                rating = '0.0'

            try:
                people_num = get_people_num(book_url)
                people_num = people_num.strip('人评价')
            except:
                people_num = '0'

            book_list.append([title, rating, people_num, author_info, pub_info])
            try_times = 0
        page_num += 1
        if page_num >= 1:
            break
        print('Downloading Information From Page %d' % page_num)
    return book_list


def do_spider(book_tag_lists):
    book_lists = []
    for book_tag in book_tag_lists:
        # get booklist by single tag and store information into a list
        book_list = book_spider(book_tag)
        # sort by rating
        book_list = sorted(book_list, key=lambda x: x[1], reverse=True)
        book_lists.append(book_list)
    return book_lists


def get_people_num(url):
    try:
        req = urllib.request.Request(url, headers=hds[np.random.randint(0, len(hds))])
        source_code = request.urlopen(req)
    except error.HTTPError as e:
        print(e)
    soup = BeautifulSoup(source_code, "lxml")
    people_num = soup.find('div', {'class': 'rating_sum'}).findAll('span')[1].string.strip()
    return people_num


def print_book_lists_excel(book_list, book_tag_lists):
    wb = Workbook()
    ws = []
    for i in range(len(book_tag_lists)):
        ws.append(wb.create_sheet(title=book_tag_lists[i]))
    for i in range(len(book_tag_lists)):
        ws[i].append(['Index', 'Name', 'Ratings', 'Rating Number', 'Author', 'Publication'])
        count = 1
        for bl in book_list[i]:
            ws[i].append([count, bl[0], float(bl[1]), int(bl[2]), bl[3], bl[4]])
            count += 1
        save_path = 'book_list'
        for i in range(len(book_tag_lists)):
            save_path += ('-' + book_tag_lists[i])
        save_path += '.xlsx'
        wb.save(save_path)


if __name__ == '__main__':
    book_tag_lists = ['linux', 'android']
    book_list = do_spider(book_tag_lists)
    print_book_lists_excel(book_list, book_tag_lists)
